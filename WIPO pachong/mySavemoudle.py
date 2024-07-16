'''
原本写的用csv的存储方式 后面改成xlsx格式的存储方式
import csv
import os

class MySaveData:
    header = None  # 类属性，用于存储表头

    def __init__(self, filename, data, header=None):
        self.filename = filename
        self.data = data
        self.data = [str(item) for item in self.data]
        if header and MySaveData.header is None:  # 如果提供了表头并且类属性header尚未设置
            MySaveData.header = header

    def csv_save(self):
        if self.data:  # 只有当有数据时才执行保存操作

            self.data = [self.data]

            with open(f'{self.filename}.csv', mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                # 检查文件是否存在，并决定是否写入表头
                if os.path.getsize(f'{self.filename}.csv') == 0:
                    writer.writerow(MySaveData.header)  # 只写入一次表头
                for row in self.data:
                    writer.writerow(row)  # 追加数据

'''




import openpyxl
import os
class MySaveData:
    header = None

    def __init__(self, filename, data, header=None):
        self.filename = filename
        self.data = data
        self.header = header

    def xlsx_save(self):
        self.data = [self.data]
        # 检查文件是否存在，如果不存在，创建新的工作簿和表头
        if not os.path.exists(f'{self.filename}.xlsx'):
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            if self.header:
                self.sheet.append(self.header)  # 写入表头
        else:
            # 加载现有的工作簿
            self.wb = openpyxl.load_workbook(f'{self.filename}.xlsx')
            self.sheet = self.wb.active

        # 检查是否需要写入表头
        if self.header and self.sheet.max_row == 0:
            self.sheet.append(self.header)

        # 写入数据行
        if self.sheet.max_row > 0:
            # 如果工作表不为空，找到最后一行的行号
            last_row = self.sheet.max_row
            # 从下一行开始写入数据
            for row_data in self.data:
                self.sheet.append(row_data)  # 追加数据
        else:
            # 如果工作表为空，写入数据
            for row_data in self.data:
                self.sheet.append(row_data)

        # 保存工作簿到文件
        self.wb.save(f'{self.filename}.xlsx')












